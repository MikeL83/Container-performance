# -*- coding: utf-8 -*-
"""

Synopsis


"""

import argparse
import os
import pprint
import subprocess
import sys
import re
from pprint import pprint
import json
import nsequencebench
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from matplotlib.ticker import MultipleLocator, FormatStrFormatter
#majorLocator = MultipleLocator(20)
#majorFormatter = FormatStrFormatter('%d')
minorLocator = MultipleLocator(5)

DB = {'gcc': {}, 'clang': {}}
CONTAINERS = ('std::vector', 'std::list', 'std::forward_list',
              'boost::vector', 'boost::list', 'boost::slist',
              'boost::stable_vector')
SIZES = (10, 100, 1000, 10000, 50000)
PERF_INST = ('')


class PostProcess:
    def __init__(self):
        self.path = os.getcwd() + '/perf_logs/'

    def parse_perf_results(self):
        # regex = re.compile(r"\s*([\d\s]+)\s*cycles")
        regex = re.compile(
            r"\s*(?P<cycles>[\d\s]+)\s*cycles.*\n"
            r"\s*(?P<instructions>[\d\s]+)\s*instructions.*\n"
            r"\s*(?P<cache_references>.+)\s*cache-references.*\n"
            r"\s*(?P<cache_misses>.+)\s*cache-misses.*\n"
            r"\s*(?P<branches>.+)\s*branches.*\n"
            r"\s*(?P<branch_misses>.+)\s*branch-misses.*\n"
            r"\s*(?P<page_faults>.+)\s*page-faults.*\n"
            r"\s*(?P<cpu_migrations>.+)\s*cpu-migrations.*\n"
            r"\s*(?P<LLC_load_misses>.+)\s*LLC-load-misses",
            re.DOTALL | re.MULTILINE)
        values = {}
        for filename in os.listdir(self.path):
            values = {}
            with open(os.path.join(self.path, filename)) as f:
                data = f.read()
                for re_match in regex.finditer(data):
                    try:
                        values['cycles'] = int(re_match.group('cycles').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['cycles'] = ''
                    try:
                        values['instructions'] = int(
                            re_match.group('instructions').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['instructions'] = ''
                    try:
                        values['cache-references'] = int(
                            re_match.group('cache_references').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['cache-references'] = ''
                    try:
                        values['cache-misses'] = int(
                            re_match.group('cache_misses').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['cache-misses'] = ''
                    try:
                        values['branches'] = int(re_match.group('branches').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['branches'] = ''
                    try:
                        values['branch-misses'] = int(
                            re_match.group('branch_misses').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['branch-misses'] = ''
                    try:
                        values['cpu-migrations'] = int(
                            re_match.group('cpu_migrations').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['cpu-migrations'] = ''
                    try:
                        values['page-faults'] = int(
                            re_match.group('page_faults').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['page-faults'] = ''
                    try:
                        values['LLC-load-misses'] = int(
                            re_match.group('LLC_load_misses').encode('ascii', 'ignore').decode().strip())
                    except:
                        values['LLC-load-misses'] = ''

                    f_name = filename.split('_')
                    f_name = [item.replace('::', '_') for item in f_name]
                    if len(f_name) == 4:
                        if not f_name[1] in DB[f_name[0]]:
                            DB[f_name[0]][f_name[1]] = {}
                        DB[f_name[0]][f_name[1]][f_name[2][1:]] = values
                    elif len(f_name) == 5:
                        if not f_name[1] + '_' + f_name[2] in DB[f_name[0]]:
                            DB[f_name[0]][f_name[1] + '_' + f_name[2]] = {}
                        DB[f_name[0]][f_name[1] + '_' + f_name[2]][f_name[3][1:]] = values

        with open('perf_results.json', 'w') as f:
            json.dump(DB, f)

    def export_to_excel(self):
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()

        # Save the file
        wb.save("sample.xlsx")



    def plot(self):

        data = ''
        path = os.getcwd()
        with open(os.path.join(path, 'benchmark_results.json')) as data_file:
            data = json.load(data_file)

            n = np.array([10, 100, 1000, 10000, 50000])

            # note that plot returns a list of lines.  The "l1, = plot" usage
            # extracts the first element of the list into l1 using tuple
            # unpacking.  So l1 is a Line2D instance, not a sequence of lines
            lines = []
            labels = []
            plt.figure(1, figsize=(13, 12), dpi=125)
            plt.legend(bbox_to_anchor=(1, 1),
                       bbox_transform=plt.gcf().transFigure)
            # number = 20
            # cmap = plt.get_cmap('gnuplot')
            # colors = [cmap(i) for i in np.linspace(0, 1, number)]
            colors = ['black', 'orangered', 'forestgreen', 'cyan', 'dodgerblue', 'purple',
                      'crimson']
            for k, cont in enumerate(CONTAINERS, start=0):
                l1 = plt.plot(n, data['c++']['gcc'][cont], label='{}_{}'.format('gcc', cont), color=colors[k])
                l2 = plt.plot(n, data['c++']['clang'][cont], label='{}_{}'.format('clang', cont), color=colors[k],
                              linestyle='-.')
                lines.append(l1)
                lines.append(l2)
            l1 = plt.plot(n, data['python']['pure'], label='{}'.format('pure python'), color='gold', linestyle=':')
            l2 = plt.plot(n, data['python']['cython'], label='{}'.format('cython'), color='red', linestyle=':')
            l3 = plt.plot(n, data['python']['numpy'], label='{}'.format('numpy'), color='lime', linestyle=':')
            lines.append(l1)
            lines.append(l2)
            lines.append(l3)
            # l2, l3 = plt.plot(t2, np.sin(2 * np.pi * t2), '--o', t1, np.log(1 + t1), '.')
            # l4, = plt.plot(t2, np.exp(-t2) * np.sin(2 * np.pi * t2), 's-.')

            # plt.legend(lines, labels) #loc='upper right', shadow=True)
            plt.legend(loc='upper left', shadow=True, fancybox=True)
            plt.xlabel('Container size (N)', fontsize=12)
            plt.ylabel('Time (s)', fontsize=12)
            plt.title('Benchmark results for different containers', fontsize=14)
            plt.minorticks_on()
            plt.grid(b=True, which='major', linestyle='-')
            figmanager = plt.get_current_fig_manager()
            figmanager.resize(*figmanager.window.maxsize())
            axes = plt.gca()
            for j in np.arange(3):
                if j == 0:
                    axes.set_ylim([0, 0.01])
                    axes.set_xlim([0, 1000])
                    plt.savefig('benchmarks_res_N1000.png', bbox_inches='tight', dpi='figure',
                                format='png')



            with open(os.path.join(path, 'perf_results.json')) as perf_file:
                perf_results = json.load(perf_file)

                # data = ((3, 1000), (10, 3), (100, 30), (500, 800), (50, 1))

                dim = 7
                w = 0.5
                dimw = w / dim

                x = np.arange(7)
                for k, size in enumerate(SIZES):
                    plt.figure(k + 2)
                    y1 = []
                    y2 = []

                    for cont in CONTAINERS:
                        y1.append(perf_results['gcc'][cont.replace('::', '_')][str(size)]['instructions'])
                        y2.append(perf_results['clang'][cont.replace('::', '_')][str(size)]['instructions'])

                    y1 = np.asarray(y1)
                    y2 = np.asarray(y2)
                    b1 = plt.bar(x + dimw / 2, y1 / y1, dimw, bottom=0.001)
                    b2 = plt.bar(x - dimw / 2, y2 / y1, dimw, bottom=0.001)

                    plt.xlabel("FOO")
                    plt.ylabel("FOO")
                    plt.title("Testing")
                    #plt.yscale('log')

                    axes = plt.gca()
                    axes.set_xticklabels(['', 'std::vector', 'std::list', 'std::forward_list',
              'boost::vector', 'boost::list', 'boost::slist',
              'boost::stable_vector'], rotation=-22.5)
                    axes.set_yticklabels([0,1,2])
                    axes.set_ylim([0, 2])
                    plt.yticks(np.array([0,1,2]))
                    #axes.yaxis.set_major_locator(majorLocator)
                    #axes.yaxis.set_major_formatter(majorFormatter)

                    # for the minor ticks, use no labels; default NullFormatter
                    axes.yaxis.set_minor_locator(minorLocator)
                    #plt.minorticks_on()
                    #plt.grid(b=True, which='major', linestyle='-')

                    # x = np.arange(7)
                    # for i in range(len(data[0])):
                    #  y = [d[i] for d in data]
                    #   b = plt.bar(x + i * dimw, y, dimw, bottom=0.001)

                    # plt.xticks(x + dimw / 2, map(str, x))

                plt.show()


if __name__ == "__main__":
    pp = PostProcess()
    pp.parse_perf_results()
    pp.export_to_excel()
    pp.plot()
